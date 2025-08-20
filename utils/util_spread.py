import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def create_spread(data, workbook_path: str, worksheet_name: str):
    def flatten_dict(d, parent_key='', sep='_'):
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(flatten_dict(v, new_key, sep=sep).items())
            elif isinstance(v, list):
                # Convert arrays to space-separated strings
                items.append((new_key, ' '.join(str(item) for item in v)))
            else:
                items.append((new_key, v))
        return dict(items)
    
    def create_nested_headers(data_list):
        if not data_list:
            return [], []
        
        all_keys = set()
        for item in data_list:
            all_keys.update(item.keys())
        
        # Use first dictionary's key order if available, otherwise sort
        if data_list:
            first_dict_keys = list(data_list[0].keys())
            # Start with first dictionary's order, then add any missing keys sorted
            ordered_keys = first_dict_keys + sorted([k for k in all_keys if k not in first_dict_keys])
            all_keys = ordered_keys
        else:
            all_keys = sorted(all_keys)
        
        # Create hierarchical headers
        header1 = []
        header2 = []
        
        for key in all_keys:
            parts = key.split('_')
            if len(parts) == 1:
                header1.append(parts[0])
                header2.append('')
            else:
                header1.append(parts[0])
                header2.append('_'.join(parts[1:]))
        
        return header1, header2
    
    # Load or create workbook
    try:
        workbook = load_workbook(workbook_path)
    except FileNotFoundError:
        workbook = Workbook()
        # Remove default sheet if it exists and we're creating a new workbook
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
    
    # Remove existing worksheet if it exists
    if worksheet_name in workbook.sheetnames:
        workbook.remove(workbook[worksheet_name])
    
    # Create new worksheet
    worksheet = workbook.create_sheet(worksheet_name)
    if not data:
        print("No data found. Returning empty workbook")
    if not isinstance(data, list):
        print("found data, but not a list")
        print(" type of data is ", type(data))
        print(json.dumps(data, indent = 2))
    
    if not data or not isinstance(data, list):
        print("No data found. Returning empty workbook")
        workbook.save(workbook_path)
        return
    
    # Flatten nested dictionaries
    flattened_data = [flatten_dict(item) if isinstance(item, dict) else item for item in data]
    
    # Create headers
    header1, header2 = create_nested_headers(flattened_data)
    
    # Header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    
    current_row = 1
    
    # Write first header row
    for col, header in enumerate(header1, 1):
        cell = worksheet.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_alignment
    
    current_row += 1
    
    # Write second header row if needed
    if any(h for h in header2):
        for col, header in enumerate(header2, 1):
            cell = worksheet.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = wrap_alignment
        current_row += 1
    
    # Write data rows
    # Use first dictionary's key order if available
    if flattened_data:
        first_dict_keys = list(flattened_data[0].keys())
        all_keys_set = set().union(*(d.keys() for d in flattened_data))
        # Start with first dictionary's order, then add any missing keys sorted
        all_keys = first_dict_keys + sorted([k for k in all_keys_set if k not in first_dict_keys])
    else:
        all_keys = [key for key in sorted(set().union(*(d.keys() for d in flattened_data)))]
    
    for item in flattened_data:
        for col, key in enumerate(all_keys, 1):
            value = item.get(key, '')
            cell = worksheet.cell(row=current_row, column=col, value=value)
            cell.alignment = wrap_alignment
        current_row += 1
    
    # Auto-size columns and apply constraints
    for col in range(1, len(all_keys) + 1):
        column_letter = get_column_letter(col)
        
        # Calculate max width based on content
        max_width = 0
        for row in range(1, current_row):
            cell_value = worksheet.cell(row=row, column=col).value
            if cell_value:
                max_width = max(max_width, len(str(cell_value)))
        
        # Set column width (max 30 chars, min based on content)
        column_width = min(max_width + 2, 30)
        worksheet.column_dimensions[column_letter].width = column_width
    
    # Enable autofilter
    if current_row > 2:  # Only if we have data beyond headers
        header_end_row = 2 if any(h for h in header2) else 1
        worksheet.auto_filter.ref = f"A{header_end_row}:{get_column_letter(len(all_keys))}{current_row - 1}"
    
    # Set freeze panes after header and first column
    if current_row > 2:  # Only if we have data
        header_end_row = 2 if any(h for h in header2) else 1
        # Freeze after the header row(s) and after the first column (B and row after header)
        freeze_cell = f"B{header_end_row + 1}"
        worksheet.freeze_panes = freeze_cell
    
    # Save workbook
    workbook.save(workbook_path)
    
